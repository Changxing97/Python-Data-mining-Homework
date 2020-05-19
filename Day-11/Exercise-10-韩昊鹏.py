from openpyxl import Workbook
import json

with open(r"D:\hhp代码文件夹\day11\聆听丶芒果鱼直播间时间切片弹幕.json", encoding="utf-8") as f:
    danmu = json.load(f)
    data = danmu["data"]

wb = Workbook()
ws1 = wb.create_sheet("Mysheet")

ws1["A1"].value = "类别"
ws1["B1"].value = "发布时间"
ws1["C1"].value = "发布者"
ws1["D1"].value = "弹幕内容"

num = 2

for i in data:
    ws1.cell(row=num, column=1).value = i[0]
    ws1.cell(row=num, column=2).value = i[1]
    ws1.cell(row=num, column=3).value = i[2]
    ws1.cell(row=num, column=4).value = i[3]
    num = num + 1

wb.save("output.xlsx")
